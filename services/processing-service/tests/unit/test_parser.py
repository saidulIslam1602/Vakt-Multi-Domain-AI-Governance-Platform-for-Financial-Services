"""Unit tests for the document parser — XLSX, TXT, DOCX, and PDF paths."""

from __future__ import annotations

import io

import pytest

from allergo_shared.domain.enums import DocumentType
from allergo_shared.domain.exceptions import ValidationError
from processing_service.infrastructure.parser import (
    _parse_text,
    _parse_xlsx,
    parse_document,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_xlsx(sheet_names_and_rows: dict) -> bytes:
    """Build a minimal in-memory XLSX file using openpyxl."""
    import openpyxl
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, rows in sheet_names_and_rows.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── _parse_text ───────────────────────────────────────────────────────────────

class TestParseText:
    def test_plain_text_extracted(self):
        data = b"Hello, world!\nSecond line."
        result = _parse_text(data)
        assert "Hello, world!" in result.text
        assert result.page_count == 1
        assert result.used_ocr is False

    def test_invalid_utf8_replaced(self):
        data = b"Valid \xff\xfe text"
        result = _parse_text(data)
        assert result.text  # Should not raise
        assert result.page_count == 1


# ── _parse_xlsx ───────────────────────────────────────────────────────────────

class TestParseXlsx:
    def test_single_sheet_content_extracted(self):
        data = _make_xlsx({"Sheet1": [["Amount", "Date"], ["1000", "2026-01-15"]]})
        result = _parse_xlsx(data)
        assert "Amount" in result.text
        assert "1000" in result.text
        assert result.page_count == 1

    def test_multiple_sheets_page_count(self):
        """page_count must equal the number of sheets (was broken before fix)."""
        data = _make_xlsx({
            "January": [["Item", "Cost"], ["Rent", "5000"]],
            "February": [["Item", "Cost"], ["Utilities", "800"]],
            "March": [["Item", "Cost"], ["Insurance", "1200"]],
        })
        result = _parse_xlsx(data)
        # This is the critical regression test — was crashing after wb.close()
        assert result.page_count == 3

    def test_page_count_after_close_does_not_raise(self):
        """The old code called wb.sheetnames after wb.close() — verify fix holds."""
        data = _make_xlsx({"Q1": [["a", "b"]], "Q2": [["c", "d"]]})
        # Should not raise AttributeError / ResourceWarning
        result = _parse_xlsx(data)
        assert result.page_count == 2

    def test_all_sheet_headers_present_in_text(self):
        data = _make_xlsx({
            "Invoices": [["Vendor", "Amount"], ["ACME", "9999"]],
            "Contracts": [["Name", "End Date"], ["Lease Oslo", "2027-12-31"]],
        })
        result = _parse_xlsx(data)
        assert "Invoices" in result.text
        assert "Contracts" in result.text
        assert "ACME" in result.text
        assert "Lease Oslo" in result.text

    def test_empty_sheet_excluded_from_text(self):
        data = _make_xlsx({"Empty": [], "HasData": [["x", "y"]]})
        result = _parse_xlsx(data)
        assert "HasData" in result.text
        # Empty sheet header should not appear since no rows
        assert "--- Sheet: Empty ---" not in result.text


# ── parse_document dispatcher ─────────────────────────────────────────────────

class TestParseDocumentDispatcher:
    def test_txt_dispatched_correctly(self):
        result = parse_document(b"plain text", DocumentType.TXT, "notes.txt")
        assert result.text == "plain text"

    def test_xlsx_dispatched_correctly(self):
        data = _make_xlsx({"Data": [["Col1", "Col2"], ["val1", "val2"]]})
        result = parse_document(data, DocumentType.XLSX, "data.xlsx")
        assert "Col1" in result.text

    def test_unknown_type_raises_validation_error(self):
        with pytest.raises(ValidationError, match="No parser available"):
            parse_document(b"data", DocumentType.UNKNOWN, "file.bin")  # type: ignore[attr-defined]
