"""Document parser — extracts raw text from PDF, DOCX, XLSX, TXT, images."""

from __future__ import annotations

import io
from dataclasses import dataclass

import fitz  # PyMuPDF
import pytesseract
from PIL import Image

from allergo_shared.domain.enums import DocumentType
from allergo_shared.domain.exceptions import ValidationError
from allergo_shared.infrastructure.logging import get_logger

logger = get_logger(__name__)


@dataclass
class ParseResult:
    text: str
    page_count: int
    used_ocr: bool = False


def parse_document(data: bytes, document_type: DocumentType, filename: str) -> ParseResult:
    """Synchronous parse — run in threadpool executor from async context."""
    if document_type == DocumentType.PDF:
        return _parse_pdf(data)
    if document_type == DocumentType.IMAGE:
        return _parse_image(data)
    if document_type in (DocumentType.TXT, DocumentType.HTML):
        return _parse_text(data)
    if document_type == DocumentType.DOCX:
        return _parse_docx(data)
    if document_type == DocumentType.XLSX:
        return _parse_xlsx(data)
    raise ValidationError(f"No parser available for document type '{document_type}'.")


def _parse_pdf(data: bytes) -> ParseResult:
    text_parts: list[str] = []
    used_ocr = False

    with fitz.open(stream=data, filetype="pdf") as doc:
        page_count = doc.page_count
        for page in doc:
            page_text = page.get_text("text").strip()
            if page_text:
                text_parts.append(page_text)
            else:
                # Fallback to OCR for image-only pages
                pix = page.get_pixmap(dpi=200)
                img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                ocr_text = pytesseract.image_to_string(img).strip()
                if ocr_text:
                    text_parts.append(ocr_text)
                    used_ocr = True

    return ParseResult(
        text="\n\n".join(text_parts),
        page_count=page_count,
        used_ocr=used_ocr,
    )


def _parse_image(data: bytes) -> ParseResult:
    img = Image.open(io.BytesIO(data))
    text = pytesseract.image_to_string(img).strip()
    return ParseResult(text=text, page_count=1, used_ocr=True)


def _parse_text(data: bytes) -> ParseResult:
    text = data.decode("utf-8", errors="replace")
    return ParseResult(text=text, page_count=1)


def _parse_docx(data: bytes) -> ParseResult:
    import docx  # python-docx — declared in pyproject.toml
    doc = docx.Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return ParseResult(text="\n".join(paragraphs), page_count=1)


def _parse_xlsx(data: bytes) -> ParseResult:
    import openpyxl  # openpyxl — declared in pyproject.toml
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    sheet_count = len(wb.sheetnames)  # capture BEFORE close()
    for sheet in wb.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"--- Sheet: {sheet.title} ---\n" + "\n".join(rows))
    wb.close()
    return ParseResult(text="\n\n".join(parts), page_count=sheet_count)
